#!/usr/bin/env python3
"""xlsx_parser.py — extract an XLSX/XLS workbook to Markdown tables.

Usage:   python3 xlsx_parser.py <file_path>
Output:  stdout — Markdown tables (UTF-8), one section per non-empty sheet
         stderr — error messages
         exit 0 success / 1 failure / 2 dependency missing

Uses openpyxl in read_only mode: a streaming reader that keeps memory
constant regardless of sheet size, so multi-megabyte worksheets parse in
seconds instead of exhausting RAM/CPU like a full DOM parse would.
"""
# VERSION: 2
import sys


def build_table(rows):
    width = max(len(r) for r in rows)
    lines = []
    for i, r in enumerate(rows):
        padded = r + [""] * (width - len(r))
        lines.append("| " + " | ".join(padded) + " |")
        if i == 0:
            lines.append("|" + " --- |" * width)
    return "\n".join(lines)


def main():
    if len(sys.argv) < 2:
        sys.stderr.write("Usage: xlsx_parser.py <file_path>\n")
        sys.exit(1)
    path = sys.argv[1]

    try:
        from openpyxl import load_workbook
    except ImportError as e:
        sys.stderr.write(f"openpyxl not installed: {e}\n")
        sys.stderr.write("Install with: pip3 install --user openpyxl\n")
        sys.exit(2)

    try:
        wb = load_workbook(path, read_only=True, data_only=True)
    except Exception as e:
        sys.stderr.write(f"Failed to open workbook: {e}\n")
        sys.exit(1)

    sections = []
    for ws in wb.worksheets:
        rows = []
        for row in ws.iter_rows(values_only=True):
            cells = ["" if v is None else str(v) for v in row]
            if any(c != "" for c in cells):
                rows.append(cells)
        if rows:
            sections.append(f"### {ws.title}\n\n{build_table(rows)}")
    wb.close()

    if not sections:
        sys.stderr.write("Spreadsheet appears to be empty\n")
        sys.exit(1)

    sys.stdout.write("\n\n".join(sections))


if __name__ == "__main__":
    main()
